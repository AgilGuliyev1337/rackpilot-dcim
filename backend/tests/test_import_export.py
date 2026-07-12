import io

from openpyxl import load_workbook

from tests.conftest import make_infra, register_org

CSV_HEADER = "name,asset_tag,serial_number,device_type,status,rack_name,position_u,height_u"


def _upload(client, headers, csv_body: str, filename: str = "import.csv"):
    return client.post(
        "/api/devices/import",
        headers=headers,
        files={"file": (filename, csv_body.encode(), "text/csv")},
    )


def _make_device(client, headers, **overrides):
    payload = {
        "name": "existing",
        "asset_tag": "AZT-0001",
        "serial_number": "SN-0001",
        "device_type": "server",
    }
    payload.update(overrides)
    r = client.post("/api/devices", json=payload, headers=headers)
    assert r.status_code == 201, r.text
    return r.json()


def test_import_valid_file(client):
    headers = register_org(client, "Imp Org", "imp@test.az")
    make_infra(client, headers)  # creates RACK-01

    body = "\n".join(
        [
            CSV_HEADER,
            "srv-a,AZT-0101,SN-0101,server,active,RACK-01,10,2",
            "sw-a,AZT-0102,SN-0102,switch,active,,,1",
        ]
    )
    r = _upload(client, headers, body)
    assert r.status_code == 200
    data = r.json()
    assert data["imported"] == 2
    assert data["failed"] == 0
    assert data["errors"] == []

    # racked device actually landed at U10-U11
    devices = client.get("/api/devices?search=srv-a", headers=headers).json()["items"]
    assert devices[0]["position_u"] == 10 and devices[0]["height_u"] == 2

    # audit log recorded the import as one action
    logs = client.get("/api/audit-logs?action=import", headers=headers).json()
    assert logs["total"] == 1
    assert logs["items"][0]["new_values"]["imported"] == 2


def test_import_bad_rows_reported_per_row(client):
    headers = register_org(client, "Bad Org", "bad@test.az")
    make_infra(client, headers)

    body = "\n".join(
        [
            CSV_HEADER,
            "ok-1,AZT-0201,SN-0201,server,active,,,1",
            ",AZT-0202,SN-0202,server,active,,,1",  # missing name
            "bad-type,AZT-0203,SN-0203,toaster,active,,,1",  # invalid enum
            "bad-rack,AZT-0204,SN-0204,server,active,NO-SUCH-RACK,1,1",  # unknown rack
            "no-fit,AZT-0205,SN-0205,server,active,RACK-01,42,4",  # exceeds 42U
        ]
    )
    r = _upload(client, headers, body)
    assert r.status_code == 200
    data = r.json()
    assert data["imported"] == 1
    assert data["failed"] == 4
    fields = {(e["row"], e["field"]) for e in data["errors"]}
    assert (3, "name") in fields
    assert (4, "device_type") in fields
    assert (5, "rack_name") in fields
    assert (6, "position_u") in fields


def test_import_duplicate_asset_tag(client):
    headers = register_org(client, "Dup Org", "dup2@test.az")
    _make_device(client, headers, asset_tag="AZT-0301", serial_number="SN-0301")

    body = "\n".join(
        [
            CSV_HEADER,
            "dup-db,AZT-0301,SN-0302,server,active,,,1",  # dup vs existing org device
            "dup-a,AZT-0303,SN-0303,server,active,,,1",
            "dup-b,AZT-0303,SN-0304,server,active,,,1",  # dup within the file
        ]
    )
    r = _upload(client, headers, body)
    data = r.json()
    assert data["imported"] == 1  # only dup-a
    assert data["failed"] == 2
    assert all(e["field"] == "asset_tag" for e in data["errors"])


def test_import_rack_overlap_rejected(client):
    headers = register_org(client, "Ovl Org", "ovl@test.az")
    infra = make_infra(client, headers)
    _make_device(
        client, headers, asset_tag="AZT-0401", serial_number="SN-0401",
        rack_id=infra["rack_id"], position_u=10, height_u=2,
    )

    body = "\n".join(
        [
            CSV_HEADER,
            "clash,AZT-0402,SN-0402,server,active,RACK-01,11,1",
        ]
    )
    data = _upload(client, headers, body).json()
    assert data["imported"] == 0
    assert data["failed"] == 1
    assert data["errors"][0]["field"] == "position_u"


def test_export_csv_respects_filters(client):
    headers = register_org(client, "Exp Org", "exp@test.az")
    _make_device(client, headers, name="srv-x", asset_tag="A1", serial_number="S1",
                 device_type="server")
    _make_device(client, headers, name="sw-x", asset_tag="A2", serial_number="S2",
                 device_type="switch")
    _make_device(client, headers, name="srv-y", asset_tag="A3", serial_number="S3",
                 device_type="server", status="inactive")

    r = client.get("/api/devices/export?format=csv&device_type=server&status=active",
                   headers=headers)
    assert r.status_code == 200
    lines = r.content.decode("utf-8-sig").strip().splitlines()
    assert len(lines) == 2  # header + srv-x only
    assert "srv-x" in lines[1]


def test_export_xlsx_and_tenant_scope(client):
    headers_a = register_org(client, "Exp A", "expa@test.az")
    headers_b = register_org(client, "Exp B", "expb@test.az")
    _make_device(client, headers_a, name="only-a", asset_tag="A1", serial_number="S1")

    r = client.get("/api/devices/export?format=xlsx", headers=headers_b)
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws.max_row == 1  # header only — org B sees none of org A's devices


def test_import_template_available(client):
    headers = register_org(client, "Tpl Org", "tpl@test.az")
    r = client.get("/api/devices/import/template?format=csv", headers=headers)
    assert r.status_code == 200
    lines = r.content.decode("utf-8-sig").strip().splitlines()
    assert lines[0].startswith("name,asset_tag,serial_number")
    assert len(lines) == 3  # header + 2 example rows


def test_import_requires_engineer(client, db):
    from app.models import UserRole
    from tests.conftest import create_user_with_role

    headers = register_org(client, "Role Imp", "roleimp@test.az")
    org_id = client.get("/api/auth/me", headers=headers).json()["organization_id"]
    viewer = create_user_with_role(db, client, org_id, "impview@test.az", UserRole.viewer)

    r = _upload(client, viewer, CSV_HEADER + "\nx,A1,S1,server,active,,,1")
    assert r.status_code == 403
