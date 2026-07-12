import csv
import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Device, DeviceStatus, DeviceType, Rack, Room, User
from app.services.audit_service import log_action

COLUMNS = [
    "name",
    "asset_tag",
    "serial_number",
    "vendor",
    "model",
    "device_type",
    "status",
    "owner",
    "notes",
    "cpu",
    "ram",
    "storage",
    "ip_address",
    "mac_address",
    "operating_system",
    "rack_name",
    "position_u",
    "height_u",
    "warranty_expiry",
]

TEMPLATE_ROWS = [
    {
        "name": "example-srv-01",
        "asset_tag": "AZT-9001",
        "serial_number": "SN-EXAMPLE-01",
        "vendor": "Dell",
        "model": "PowerEdge R750",
        "device_type": "server",
        "status": "active",
        "owner": "Platform Team",
        "notes": "",
        "cpu": "2x Intel Xeon Gold 6338",
        "ram": "256GB DDR4",
        "storage": "8x 1.92TB SSD",
        "ip_address": "10.10.5.11",
        "mac_address": "00:1B:44:99:01:A0",
        "operating_system": "Ubuntu Server 22.04 LTS",
        "rack_name": "RACK-01",
        "position_u": "20",
        "height_u": "2",
        "warranty_expiry": "2027-06-30",
    },
    {
        "name": "example-sw-01",
        "asset_tag": "AZT-9002",
        "serial_number": "SN-EXAMPLE-02",
        "vendor": "Cisco",
        "model": "Catalyst 9300-48T",
        "device_type": "switch",
        "status": "active",
        "owner": "Network Team",
        "notes": "unracked example",
        "cpu": "",
        "ram": "",
        "storage": "",
        "ip_address": "10.10.5.12",
        "mac_address": "",
        "operating_system": "",
        "rack_name": "",
        "position_u": "",
        "height_u": "1",
        "warranty_expiry": "",
    },
]


class ImportError_(BaseModel):
    row: int
    field: str
    message: str


class ImportResult(BaseModel):
    imported: int
    failed: int
    errors: list[ImportError_]


def _device_row(device: Device, rack_names: dict[int, str]) -> dict[str, Any]:
    return {
        "name": device.name,
        "asset_tag": device.asset_tag,
        "serial_number": device.serial_number,
        "vendor": device.vendor or "",
        "model": device.model or "",
        "device_type": device.device_type.value,
        "status": device.status.value,
        "owner": device.owner or "",
        "notes": device.notes or "",
        "cpu": device.cpu or "",
        "ram": device.ram or "",
        "storage": device.storage or "",
        "ip_address": device.ip_address or "",
        "mac_address": device.mac_address or "",
        "operating_system": device.operating_system or "",
        "rack_name": rack_names.get(device.rack_id, "") if device.rack_id else "",
        "position_u": device.position_u if device.position_u is not None else "",
        "height_u": device.height_u,
        "warranty_expiry": (
            device.warranty_expiry.isoformat() if device.warranty_expiry else ""
        ),
    }


def _rows_to_csv(rows: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def _rows_to_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_devices(
    db: Session,
    org_id: int,
    *,
    fmt: str,
    device_type: DeviceType | None = None,
    status: DeviceStatus | None = None,
    datacenter_id: int | None = None,
) -> tuple[bytes, str, str]:
    """Returns (content, media_type, filename)."""
    query = select(Device).where(Device.organization_id == org_id).order_by(Device.name)
    if device_type is not None:
        query = query.where(Device.device_type == device_type)
    if status is not None:
        query = query.where(Device.status == status)
    if datacenter_id is not None:
        query = (
            query.join(Rack, Device.rack_id == Rack.id)
            .join(Room, Rack.room_id == Room.id)
            .where(Room.datacenter_id == datacenter_id)
        )
    devices = list(db.scalars(query))
    rack_names = {
        r.id: r.name
        for r in db.scalars(select(Rack).where(Rack.organization_id == org_id))
    }
    rows = [_device_row(d, rack_names) for d in devices]

    stamp = date.today().isoformat()
    if fmt == "xlsx":
        return (
            _rows_to_xlsx(rows),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"devices-{stamp}.xlsx",
        )
    return _rows_to_csv(rows), "text/csv; charset=utf-8", f"devices-{stamp}.csv"


def build_template(fmt: str) -> tuple[bytes, str, str]:
    if fmt == "xlsx":
        return (
            _rows_to_xlsx(TEMPLATE_ROWS),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "devices-import-template.xlsx",
        )
    return (
        _rows_to_csv(TEMPLATE_ROWS),
        "text/csv; charset=utf-8",
        "devices-import-template.csv",
    )


def _parse_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    """Parse CSV or XLSX into a list of row dicts keyed by header."""
    if filename.lower().endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        return [
            {
                header[i]: (str(cell).strip() if cell is not None else "")
                for i, cell in enumerate(row)
                if i < len(header)
            }
            for row in rows_iter
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
    # CSV (utf-8 with optional BOM)
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [
        {(k or "").strip(): (v or "").strip() for k, v in row.items()}
        for row in reader
        if any((v or "").strip() for v in row.values())
    ]


def import_devices(
    db: Session, user: User, filename: str, content: bytes
) -> ImportResult:
    org_id = user.organization_id
    rows = _parse_upload(filename, content)
    errors: list[ImportError_] = []
    imported = 0

    racks = {
        r.name: r
        for r in db.scalars(select(Rack).where(Rack.organization_id == org_id))
    }
    existing_tags = set(
        db.scalars(select(Device.asset_tag).where(Device.organization_id == org_id))
    )
    existing_serials = set(
        db.scalars(select(Device.serial_number).where(Device.organization_id == org_id))
    )
    # occupancy per rack id: set of occupied U
    occupancy: dict[int, set[int]] = {}
    for d in db.scalars(
        select(Device).where(
            Device.organization_id == org_id,
            Device.rack_id.is_not(None),
            Device.position_u.is_not(None),
        )
    ):
        occ = occupancy.setdefault(d.rack_id, set())
        occ.update(range(d.position_u, d.position_u + d.height_u))

    valid_types = {t.value for t in DeviceType}
    valid_statuses = {s.value for s in DeviceStatus}

    for idx, row in enumerate(rows, start=2):  # row 1 is the header
        row_errors: list[ImportError_] = []

        def err(field: str, message: str) -> None:
            row_errors.append(ImportError_(row=idx, field=field, message=message))

        name = row.get("name", "")
        asset_tag = row.get("asset_tag", "")
        serial = row.get("serial_number", "")
        dtype = row.get("device_type", "")
        status_s = row.get("status", "") or "active"

        if not name:
            err("name", "Required")
        if not asset_tag:
            err("asset_tag", "Required")
        elif asset_tag in existing_tags:
            err("asset_tag", f"Duplicate asset_tag '{asset_tag}' in organization")
        if not serial:
            err("serial_number", "Required")
        elif serial in existing_serials:
            err("serial_number", f"Duplicate serial_number '{serial}' in organization")
        if dtype not in valid_types:
            err("device_type", f"Invalid type '{dtype}' (valid: {', '.join(sorted(valid_types))})")
        if status_s not in valid_statuses:
            err("status", f"Invalid status '{status_s}'")

        height_u = 1
        if row.get("height_u"):
            try:
                height_u = max(1, int(float(row["height_u"])))
            except ValueError:
                err("height_u", f"Not a number: '{row['height_u']}'")

        warranty: date | None = None
        if row.get("warranty_expiry"):
            raw = row["warranty_expiry"].split(" ")[0]
            try:
                warranty = datetime.strptime(raw, "%Y-%m-%d").date()
            except ValueError:
                err("warranty_expiry", f"Invalid date '{row['warranty_expiry']}' (use YYYY-MM-DD)")

        rack = None
        position_u: int | None = None
        rack_name = row.get("rack_name", "")
        if rack_name:
            rack = racks.get(rack_name)
            if not rack:
                err("rack_name", f"Rack '{rack_name}' not found in organization")
            elif not row.get("position_u"):
                err("position_u", "Required when rack_name is set")
            else:
                try:
                    position_u = int(float(row["position_u"]))
                except ValueError:
                    err("position_u", f"Not a number: '{row['position_u']}'")
                if rack and position_u is not None:
                    top = position_u + height_u - 1
                    if position_u < 1 or top > rack.u_height:
                        err(
                            "position_u",
                            f"U{position_u}-U{top} does not fit in {rack.u_height}U rack",
                        )
                    else:
                        occ = occupancy.setdefault(rack.id, set())
                        span = set(range(position_u, position_u + height_u))
                        if occ & span:
                            err("position_u", f"Overlaps an occupied slot in '{rack_name}'")

        if row_errors:
            errors.extend(row_errors)
            continue

        device = Device(
            name=name,
            asset_tag=asset_tag,
            serial_number=serial,
            vendor=row.get("vendor") or None,
            model=row.get("model") or None,
            device_type=DeviceType(dtype),
            status=DeviceStatus(status_s),
            owner=row.get("owner") or None,
            notes=row.get("notes") or None,
            cpu=row.get("cpu") or None,
            ram=row.get("ram") or None,
            storage=row.get("storage") or None,
            ip_address=row.get("ip_address") or None,
            mac_address=row.get("mac_address") or None,
            operating_system=row.get("operating_system") or None,
            rack_id=rack.id if rack else None,
            position_u=position_u,
            height_u=height_u,
            warranty_expiry=warranty,
            organization_id=org_id,
        )
        db.add(device)
        imported += 1
        existing_tags.add(asset_tag)
        existing_serials.add(serial)
        if rack and position_u is not None:
            occupancy.setdefault(rack.id, set()).update(
                range(position_u, position_u + height_u)
            )

    if imported:
        db.flush()
    log_action(
        db,
        user=user,
        action="import",
        entity_type="device",
        entity_id=0,
        entity_name=f"{imported} devices from {filename}",
        new_values={"imported": imported, "failed": len({e.row for e in errors})},
    )
    db.commit()
    failed_rows = len({e.row for e in errors})
    return ImportResult(imported=imported, failed=failed_rows, errors=errors)
